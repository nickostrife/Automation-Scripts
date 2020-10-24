import openpyxl
import os
from pathlib import Path

wb = openpyxl.load_workbook(Path('C:/Users/Nicko/Documents/automate_online-materials/example.xlsx'))
# print(wb.sheetnames)
# sheet = wb['Sheet3']
# print(sheet)
# print(type(sheet))
# print(sheet.title)
# anotherSheet = wb.active
# print(anotherSheet)

sheet = wb['Sheet1']
print("sheet['A1'] = {}".format(sheet['A1']))
print("sheet['A1'].value = {}".format(sheet['A1'].value))
c = sheet['B1']
print("c.value = {}".format(c.value))
print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))
print("Cell {} is {}".format(c.coordinate, c.value))
print("sheet['C1'].value = {}".format(sheet['C1'].value))
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)